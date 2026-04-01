"""
Servico de Geracao de Orcamentos - Rogama SL
"""
from flask import Flask, request, jsonify, send_from_directory
import os, shutil, base64, subprocess, uuid, json
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__)

TEMPLATES = {
    'ROGAMA':   '/app/templates/Ppto Rogama - 2022.xlsx',
    'MULTIMAP': '/app/templates/Presupuesto Multimap Logo Nuevo.xlsm'
}

FILES_DIR = '/app/files'
os.makedirs(FILES_DIR, exist_ok=True)
BASE_URL = os.environ.get('BASE_URL', 'https://n8n-rogama-orcamentos.ht493o.easypanel.host')


def calcular_cantidad(item):
    mediciones = item.get('mediciones', [])
    if mediciones:
        total = 0
        for m in mediciones:
            parcial = m.get('parcial')
            if parcial:
                total += parcial
            else:
                uds = m.get('uds', 1) or 1
                lon = m.get('longitud', 1) or 1
                anc = m.get('anchura', 1) or 1
                alt = m.get('altura', 1) or 1
                total += uds * lon * anc * alt
        return round(total, 3)
    return item.get('cantidad', 1)


def preencher_rogama(orcamento, dest):
    wb = load_workbook(dest)
    ws = wb['PRESUPUESTO']
    ws['F42'] = orcamento.get('expediente', '')
    ws['F43'] = orcamento.get('cliente', '')
    ws['F44'] = orcamento.get('direccion', '')
    ws['F45'] = orcamento.get('localidad', '')
    ws['F46'] = orcamento.get('cp', '')
    ws['F47'] = orcamento.get('telefono', '')
    ws['F48'] = orcamento.get('fecha', datetime.now().strftime('%d/%m/%Y'))
    ws['I64'] = orcamento.get('expediente', '')
    ws['I65'] = f"{orcamento.get('direccion', '')} - {orcamento.get('localidad', '')}"

    items = orcamento.get('items', [])
    total_sem_iva = 0

    for i, item in enumerate(items[:11]):
        row = 71 + i
        qty = calcular_cantidad(item)
        precio = item.get('precio_unitario', 0)
        ws.cell(row=row, column=1).value = item.get('codigo', '')
        ws.cell(row=row, column=5).value = item.get('concepto_corto') or item.get('concepto', '')
        ws.cell(row=row, column=10).value = qty
        ws.cell(row=row, column=11).value = precio
        total_sem_iva += qty * precio

    iva = round(total_sem_iva * 0.21, 2)
    ws['L82'] = round(total_sem_iva, 2)
    ws['K83'] = 0.21
    ws['L83'] = iva
    ws['L84'] = round(total_sem_iva + iva, 2)
    ws['J86'] = datetime.now().strftime('%d/%m/%Y')
    wb.save(dest)


def preencher_multimap(orcamento, dest):
    wb = load_workbook(dest, keep_vba=True)
    ws = wb['PRESUPUESTO']
    ws['E5'] = orcamento.get('expediente', '')
    ws['E6'] = orcamento.get('cliente', '')
    ws['E7'] = orcamento.get('direccion', '')
    ws['E8'] = orcamento.get('localidad', '')
    ws['E9'] = orcamento.get('cp', '')
    ws['E10'] = orcamento.get('fecha', datetime.now().strftime('%d/%m/%Y'))
    ws['E11'] = orcamento.get('telefono', '')

    items = orcamento.get('items', [])
    for i, item in enumerate(items[:466]):
        row = 136 + i
        codigo = item.get('codigo', '')
        concepto = item.get('concepto_corto') or item.get('concepto', '')
        unidad = item.get('unidad', '')
        if not unidad:
            c = concepto.lower()
            if 'm2' in c or 'm2' in codigo.lower(): unidad = 'm2'
            elif 'ml' in c or 'ml' in codigo.lower(): unidad = 'ml'
            elif 'm3' in c or 'm3' in codigo.lower(): unidad = 'm3'
            else: unidad = 'ud'
        ws.cell(row=row, column=2).value = unidad
        ws.cell(row=row, column=3).value = concepto
        ws.cell(row=row, column=4).value = calcular_cantidad(item)
        ws.cell(row=row, column=5).value = item.get('precio_unitario', 0)
    wb.save(dest)


def excel_para_pdf(excel_path):
    try:
        output_dir = os.path.dirname(excel_path)
        subprocess.run(['libreoffice', '--headless', '--convert-to', 'pdf',
                        '--outdir', output_dir, excel_path],
                       capture_output=True, text=True, timeout=60)
        pdf_path = excel_path.rsplit('.', 1)[0] + '.pdf'
        return pdf_path if os.path.exists(pdf_path) else None
    except Exception:
        return None


@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'service': 'rogama-orcamentos'})


@app.route('/files/<filename>', methods=['GET'])
def serve_file(filename):
    return send_from_directory(FILES_DIR, filename)


@app.route('/gerar-orcamento', methods=['POST'])
def gerar_orcamento():
    try:
        data = request.get_json(force=True)
        if isinstance(data, str):
            data = json.loads(data)

        orcamento = data.get('orcamento', data)
        if isinstance(orcamento, str):
            orcamento = json.loads(orcamento)

        exp = orcamento.get('expediente', 'SEM_EXP')
        template = orcamento.get('template', '').upper()
        if not template or template not in TEMPLATES:
            template = 'MULTIMAP' if str(exp).upper().startswith('A') else 'ROGAMA'

        localidad = orcamento.get('localidad', '').replace('/', '-')[:20]
        nome_base = f"{exp} - {localidad}"
        nome_base_safe = nome_base.replace(' ', '_')
        ext = '.xlsm' if template == 'MULTIMAP' else '.xlsx'

        file_id = str(uuid.uuid4())[:8]
        excel_filename = f"{file_id}_{nome_base_safe}{ext}"
        excel_dest = os.path.join(FILES_DIR, excel_filename)
        shutil.copy(TEMPLATES[template], excel_dest)

        if template == 'MULTIMAP':
            preencher_multimap(orcamento, excel_dest)
        else:
            preencher_rogama(orcamento, excel_dest)

        excel_url = f"{BASE_URL}/files/{excel_filename}"
        pdf_url = None
        pdf_path = excel_para_pdf(excel_dest)
        if pdf_path and os.path.exists(pdf_path):
            pdf_filename = os.path.basename(pdf_path)
            pdf_url = f"{BASE_URL}/files/{pdf_filename}"
            nome_ficheiro = f"{nome_base}.pdf"
        else:
            nome_ficheiro = f"{nome_base}{ext}"

        with open(excel_dest, 'rb') as f:
            excel_b64 = base64.b64encode(f.read()).decode('utf-8')

        return jsonify({
            'success': True,
            'expediente': exp,
            'template': template,
            'nome_ficheiro': nome_ficheiro,
            'excel_url': excel_url,
            'pdf_url': pdf_url,
            'excel_base64': excel_b64,
            'pdf_base64': None
        })

    except Exception as e:
        return jsonify({'success': False, 'erro': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 80))
    app.run(host='0.0.0.0', port=port, debug=False)
